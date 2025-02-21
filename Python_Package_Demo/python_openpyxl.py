'''
Created on 13-Jul-2019

@author: siddharthsahu
'''
from openpyxl.chart import BarChart, Reference

import openpyxl as xl


def process_xl_workbook(filename):
    try:
        wb = xl.load_workbook(filename)
        sheet = wb['Sheet1']
        # cell = sheet['A2'] #to access the shell
        # cell = sheet.cell(1, 1) #to access the shell

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            corrected_price = cell.value * 0.9
            corrected_price_cell = sheet.cell(row, 4)
            corrected_price_cell.value = corrected_price

        chart_values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

        barchart = BarChart()
        barchart.add_data(chart_values)

        sheet.add_chart(barchart, 'E2')

        wb.save(f'new_{filename}')

        return True
    except:
        return False


print(process_xl_workbook('transactions.xlsx'))
